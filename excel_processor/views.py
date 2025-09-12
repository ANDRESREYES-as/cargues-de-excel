from django.views.decorators.http import require_GET
from django.utils.dateparse import parse_datetime
from django.http import HttpResponse, JsonResponse
import io
import openpyxl
import datetime
from django.core.exceptions import ObjectDoesNotExist
from .models import RegistroExcel, ExcelProcess
import traceback
# Endpoint para exportar informe histórico o filtrado por rango de fechas en Excel
@require_GET
def exportar_excel_historico(request):
    """
    Exporta un Excel con el histórico de todos los registros o filtrado por parámetros.
    Parámetros GET:
        - fecha_inicio: (opcional) formato 'YYYY-MM-DD'
        - fecha_fin: (opcional) formato 'YYYY-MM-DD'
        - op: (opcional) filtrar por número de orden de producción
    """
    try:
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        # Validar formatos de fecha
        if fecha_inicio:
            try:
                parse_datetime(fecha_inicio)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha_inicio inválido. Use YYYY-MM-DD'}, status=400)
                
        if fecha_fin:
            try:
                parse_datetime(fecha_fin)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha_fin inválido. Use YYYY-MM-DD'}, status=400)

        # Obtener registros
        qs = RegistroExcel.objects.all().select_related('proceso').order_by('id')
        
        # Aplicar filtros
        if fecha_inicio:
            qs = qs.filter(fecha_registro__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha_registro__lte=fecha_fin)
            
        # Filtro por orden de producción
        orden_produccion = request.GET.get('op')
        if orden_produccion:
            # Buscar coincidencia parcial e ignorar mayúsculas/minúsculas
            qs = qs.filter(orden__icontains=orden_produccion)

        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Histórico'
    except Exception as e:
        print(f"Error al preparar el reporte: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'error': 'Error al generar el reporte. Por favor contacte al administrador.'
        }, status=500)
    try:
        # Encabezados
        ws.append(['Archivo', 'Consecutivo', 'Orden', 'Producción', 'Cant. Orig', 
                  'Saldo Entregar', 'Cant. Produc', 'Fecha Registro'])
        
        # Datos
        for obj in qs:
            try:
                row = [
                    str(obj.proceso.archivo) if obj.proceso else '',
                    obj.proceso.consecutivo if obj.proceso else '',
                    obj.orden or '',
                    obj.produccion or '',
                    obj.cant_orig or '',
                    obj.saldo_entregar or '',
                    obj.cant_produc or '',
                    obj.fecha_registro.strftime('%Y-%m-%d %H:%M') if obj.fecha_registro else ''
                ]
                ws.append(row)
            except Exception as e:
                print(f"Error al procesar registro {obj.id}: {str(e)}")
                continue

        # Ajustar anchos de columna
        for idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column = openpyxl.utils.get_column_letter(idx)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Generar archivo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Generar nombre del archivo con los filtros aplicados
        filename_parts = ['informe_historico']
        if orden_produccion:
            filename_parts.append(f'OP_{orden_produccion}')
        if fecha_inicio:
            filename_parts.append(f'desde_{fecha_inicio}')
        if fecha_fin:
            filename_parts.append(f'hasta_{fecha_fin}')
        filename_parts.append(datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
        
        filename = '_'.join(filename_parts) + '.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        print(f"Error al generar el archivo Excel: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'error': 'Error al generar el archivo Excel. Por favor contacte al administrador.'
        }, status=500)

from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import ExcelProcess, RegistroExcel
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
from django.core.paginator import Paginator

@csrf_exempt
def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        # Filtrar filas para cada PDF
        doc_iny_3_7 = []
        doc_iny_otros = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                nuevo_val = float(row[9]) if row[9] is not None else 0
            except (ValueError, TypeError):
                nuevo_val = 0
            if nuevo_val > 0:
                try:
                    iny_val = int(row[10]) if row[10] is not None else 0
                except (ValueError, TypeError):
                    iny_val = 0
                if iny_val in (3, 7):
                    doc_iny_3_7.append(row)
                else:
                    doc_iny_otros.append(row)
        ultimo = ExcelProcess.objects.order_by('-consecutivo').first()
        consecutivo_3_7 = (ultimo.consecutivo + 1) if ultimo else 1
        excel_obj_3_7 = ExcelProcess.objects.create(archivo=archivo, consecutivo=consecutivo_3_7)
        for row in doc_iny_3_7:
            RegistroExcel.objects.create(
                proceso=excel_obj_3_7,
                orden=row[0],
                produccion=row[1],
                cant_orig=row[2],
                saldo_entregar=row[3],
                cant_produc=row[9],
                iny=row[10] if len(row) > 10 else '',
                otros='' # puedes mapear más campos si lo deseas
            )
        ultimo = ExcelProcess.objects.order_by('-consecutivo').first()
        consecutivo_otros = (ultimo.consecutivo + 1) if ultimo else consecutivo_3_7 + 1
        excel_obj_otros = ExcelProcess.objects.create(archivo=archivo, consecutivo=consecutivo_otros)
        for row in doc_iny_otros:
            RegistroExcel.objects.create(
                proceso=excel_obj_otros,
                orden=row[0],
                produccion=row[1],
                cant_orig=row[2],
                saldo_entregar=row[3],
                cant_produc=row[9],
                iny=row[10] if len(row) > 10 else '',
                otros='' # puedes mapear más campos si lo deseas
            )

        pdf_3_7 = generar_pdf('planilla', doc_iny_3_7, consecutivo_3_7)
        pdf_otros = generar_pdf('planilla', doc_iny_otros, consecutivo_otros)

        return JsonResponse({
            'pdf_3_7': pdf_3_7,
            'pdf_otros': pdf_otros,
            'consecutivo_3_7': consecutivo_3_7,
            'consecutivo_otros': consecutivo_otros
        })
    return JsonResponse({'error': 'Método no permitido o archivo no enviado'}, status=400)

def home(request):
    return render(request, 'excel_processor/home.html')

def upload_excel_web(request):
    pdf_3_7 = pdf_otros = None
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        doc_iny_3_7 = []
        doc_iny_otros = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                nuevo_val = float(row[9]) if row[9] is not None else 0
            except (ValueError, TypeError):
                nuevo_val = 0
            if nuevo_val > 0:
                try:
                    iny_val = int(row[10]) if row[10] is not None else 0
                except (ValueError, TypeError):
                    iny_val = 0
                if iny_val in (3, 7):
                    doc_iny_3_7.append(row)
                else:
                    doc_iny_otros.append(row)
        from .models import ExcelProcess, RegistroExcel
        ultimo = ExcelProcess.objects.order_by('-consecutivo').first()
        consecutivo_3_7 = (ultimo.consecutivo + 1) if ultimo else 1
        excel_obj_3_7 = ExcelProcess.objects.create(archivo=archivo, consecutivo=consecutivo_3_7)
        for row in doc_iny_3_7:
            RegistroExcel.objects.create(
                proceso=excel_obj_3_7,
                orden=row[0],
                produccion=row[1],
                cant_orig=row[2],
                saldo_entregar=row[3],
                cant_produc=row[9],
                iny=row[10] if len(row) > 10 else '',
                otros=''
            )
        ultimo = ExcelProcess.objects.order_by('-consecutivo').first()
        consecutivo_otros = (ultimo.consecutivo + 1) if ultimo else consecutivo_3_7 + 1
        excel_obj_otros = ExcelProcess.objects.create(archivo=archivo, consecutivo=consecutivo_otros)
        for row in doc_iny_otros:
            RegistroExcel.objects.create(
                proceso=excel_obj_otros,
                orden=row[0],
                produccion=row[1],
                cant_orig=row[2],
                saldo_entregar=row[3],
                cant_produc=row[9],
                iny=row[10] if len(row) > 10 else '',
                otros=''
            )
        # Generar PDFs usando la función existente
        import os
        from django.conf import settings
        pdf_3_7_path = generar_pdf('planilla', doc_iny_3_7, consecutivo_3_7)
        pdf_otros_path = generar_pdf('planilla', doc_iny_otros, consecutivo_otros)
        # Convertir rutas absolutas a rutas relativas para MEDIA_URL
        def rel_path(abs_path):
            media_root = os.path.abspath(settings.MEDIA_ROOT)
            if abs_path.startswith(media_root):
                return settings.MEDIA_URL + abs_path[len(media_root):].replace('\\','/').lstrip('/')
            return abs_path
        pdf_3_7 = rel_path(pdf_3_7_path) if pdf_3_7_path else None
        pdf_otros = rel_path(pdf_otros_path) if pdf_otros_path else None
    return render(request, 'excel_processor/upload.html', {'pdf_3_7': pdf_3_7, 'pdf_otros': pdf_otros})

def historico(request):
    # Obtener parámetros de filtrado
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    orden_produccion = request.GET.get('op')
    
    # Iniciar queryset con select_related para optimizar
    registros = RegistroExcel.objects.all().select_related('proceso').order_by('-fecha_registro')
    
    # Aplicar filtros
    if fecha_inicio:
        registros = registros.filter(fecha_registro__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha_registro__lte=fecha_fin)
    if orden_produccion:
        registros = registros.filter(orden__icontains=orden_produccion)
    paginator = Paginator(registros, 50)
    page = request.GET.get('page')
    registros_page = paginator.get_page(page)
    return render(request, 'excel_processor/historico.html', {'registros': registros_page})

def generar_pdf(nombre, filas, consecutivo):
    import datetime
    import os
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    width, height = letter
    margen_cm = 1.7
    margen = int(margen_cm * 28.3465)  # 1 cm = 28.3465 puntos
    y = height - margen
    fecha = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    encabezado = [
        'ORDEN',
        'PRODUC.',
        'CANT.\nORIG',
        'SALDO P\nENTREGAR',
        'CANT.\nPRODUC',
        'ENTREGA',
        'FALTA\nNTES'
    ]
    col_widths = [50, 140, 60, 80, 70, 60, 60]
    col_positions = [margen]
    for w in col_widths[:-1]:
        col_positions.append(col_positions[-1] + w)
    pdf_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'media')
    os.makedirs(pdf_dir, exist_ok=True)
    pdf_path = os.path.abspath(os.path.join(pdf_dir, f'{nombre}_{consecutivo}.pdf'))
    c = canvas.Canvas(pdf_path, pagesize=letter)
    def dibujar_encabezado(y, pagina):
        c.setFont('Helvetica-Bold', 18)
        c.drawCentredString(width // 2, y, "PLANILLA DE PRODUCCION")
        y -= 32
        c.setFont('Helvetica-Bold', 14)
        c.drawString(margen, y, f"Reporte consecutivo: {consecutivo}")
        c.drawRightString(width - margen, y, f"Fecha: {fecha}")
        y -= 27
        c.setFont('Helvetica-Bold', 12)
        max_lines = 1
        for col in encabezado:
            max_lines = max(max_lines, len(col.split('\n')))
        for i, col in enumerate(encabezado):
            if '\n' in col:
                lineas = col.split('\n')
                y_offset = y
                for idx, linea in enumerate(lineas):
                    c.drawCentredString(col_positions[i] + col_widths[i] // 2, y_offset - (idx * 12), linea)
            else:
                c.drawString(col_positions[i], y, str(col))
        c.setFont('Helvetica-Bold', 12)
        c.drawRightString(width - margen, height - margen + 5, f"Página {pagina}")
        y_line_top = y + 12
        y_line_bottom = y - 12 - 22
        for x in col_positions:
            c.setLineWidth(1)
            c.line(x, y_line_top, x, y_line_bottom)
        c.line(col_positions[-1] + col_widths[-1], y_line_top, col_positions[-1] + col_widths[-1], y_line_bottom)
        c.setLineWidth(1)
        c.line(margen, y_line_bottom, col_positions[-1] + col_widths[-1], y_line_bottom)
        y -= 12 + 22 + 15  # AGREGADO: 15 puntos adicionales de separación
        return y
    pagina = 1
    y = dibujar_encabezado(y, pagina)
    suma_nuevo = 0
    c.setFont('Helvetica', 11)
    for fila in filas:
        valores = [fila[0], fila[1], fila[2], fila[3], fila[9], '', '']
        try:
            suma_nuevo += float(fila[9]) if fila[9] is not None else 0
        except (ValueError, TypeError):
            pass
        for i, val in enumerate(valores):
            if i in [2, 3, 4]:
                col_center = col_positions[i] + col_widths[i] // 2
                c.drawCentredString(col_center, y, str(val))
            else:
                c.drawString(col_positions[i], y, str(val))
        y_line_top = y + 8
        y_line_bottom = y - 8
        for x in col_positions:
            c.setLineWidth(0.5)
            c.line(x, y_line_top, x, y_line_bottom)
        c.line(col_positions[-1] + col_widths[-1], y_line_top, col_positions[-1] + col_widths[-1], y_line_bottom)
        c.setLineWidth(0.5)
        c.line(margen, y_line_bottom, col_positions[-1] + col_widths[-1], y_line_bottom)
        y -= 17
        if y < 40:
            c.showPage()
            pagina += 1
            y = height - margen
            y = dibujar_encabezado(y, pagina)
            c.setFont('Helvetica', 11)
    y -= 22
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margen, y, f"Total Cantidad Producida: {suma_nuevo}")
    y -= 44
    pie_lines = [
        ("Firma Responsable: _____________________________", 27),
        ("Firma Quien recibe: ______________________________", 27),
        ("Verificado seguridad: ______________________________", 27),
        ("Comentarios: ___________________________________________________________", 17),
        ("______________________________________________________________________", 17),
        ("______________________________________________________________________", 17),
        ("", 17),
        ("Gerencia De produccion: _____________________________", 17),
        ("", 17),
        ("Gerencia General o Administrativa: _____________________________", 17),
    ]
    espacio_pie = sum([line[1] for line in pie_lines]) + 12
    if y - espacio_pie < 20:
        c.showPage()
        pagina += 1
        y = height - margen
        y = dibujar_encabezado(y, pagina)
        c.setFont('Helvetica', 11)
    c.setFont('Helvetica', 12)
    for texto, salto in pie_lines:
        c.drawString(margen, y, texto)
        y -= salto
    c.save()
    return pdf_path

import os

def pdf_list(request):
    consecutivo = request.GET.get('consecutivo')
    procesos = ExcelProcess.objects.all().order_by('-fecha_carga')
    if consecutivo:
        procesos = procesos.filter(consecutivo=consecutivo)
    from django.conf import settings
    for proc in procesos:
        pdf_name = f"planilla_{proc.consecutivo}.pdf"
        pdf_path = os.path.join('media', pdf_name)
        abs_pdf_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', pdf_path)
        if os.path.exists(abs_pdf_path):
            proc.pdf_path = f"/{pdf_path.replace('\\','/')}"
        else:
            proc.pdf_path = None
    return render(request, 'excel_processor/pdf_list.html', {'procesos': procesos})
